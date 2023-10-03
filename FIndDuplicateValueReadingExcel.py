from openpyxl import load_workbook

def are_values_equal(row, start_col, end_col):
    for i in range(start_col, end_col + 1):
        if row[i].value != row[start_col].value:
            return False
    return True

def find_rows_with_same_values(data, start_col, end_col):
    result = []
    for row_num, row in enumerate(data, start=1):
        if are_values_equal(row, start_col, end_col):
            result.append(row_num)
    return result

# Replace 'sample.xlsx' with your actual file path
file_path = 'sample.xlsx'

# Specify the columns F to AI (1-based index)
start_column = 6  # Column F
end_column = 35   # Column AI

# Load the Excel file
workbook = load_workbook(filename=file_path)
sheet = workbook.active

# Find row numbers with the same values in the specified columns
result_row_numbers = find_rows_with_same_values(sheet.iter_rows(min_row=1, values_only=True),
                                                start_column - 1, end_column - 1)

# Print the result
if len(result_row_numbers) > 0:
    print("Row numbers with the same values from column F to AI:")
    print(result_row_numbers)
else:
    print("No rows with the same values from column F to AI.")
